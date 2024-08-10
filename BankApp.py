import tkinter
from tkinter import *
import pathlib
from openpyxl import Workbook
from tkinter import  messagebox
import openpyxl, xlrd


background = "skyblue"
frame = "#fff"
fore = "skyblue"


class MyGUI:
    def __init__(self):
        self.root = Tk()
        self.root.title("Chase Bank")
        self.root.geometry("1440x900")
        self.root.config(bg=background)


        # Database
        file = pathlib.Path('Bank_Access.xlsx')
        if file.exists():
            pass

        else:
            file = Workbook()
            sheet = file.active
            sheet['A1']="Username"
            sheet['B1']="Password"

            file.save('Bank_Access.xlsx')

        
        def signup():
            F1 = Username.get()
            M1 = password.get()

            if F1=="" or M1 =="":
                messagebox.showerror("Oh! Sorry", "Few Data is missing!")


            else:
                file = openpyxl.load_workbook('Bank_Access.xlsx')
                sheet = file.active
                sheet.cell(column=1, row=sheet.max_row+1, value=F1)
                sheet.cell(column=2, row=sheet.max_row, value=M1)


                file.save('Bank_Access.xlsx')

                messagebox.showinfo("Success", "Successfully data entered")

                clear()

        def clear():
            Username.set('')
            password.set('')
            




        # Background Image
        img=PhotoImage(file="Images\seaside.png")
        lbl=Label(self.root, image=img)
        lbl.place(x=0, y=0)


        # Form Design
        form = LabelFrame(self.root, text="Registration App", font=20, bg=frame, fg=fore, width=1200, height=400)
        form.place(x=80, y=350)

        #LABELS
        Label(form, text="Enter your username", font="Arial 13",bg=frame).place(x=40, y=30)
        Label(form, text="Enter your password", font="Arial 13", bg=frame).place(x=40, y=150)


        # Entries
        Username=StringVar()
        UserName_entry = Entry(form, textvariable=Username, width=120, font="Arial 13")
        UserName_entry.place(x=40, y=80)

        password = StringVar()
        passwd_entry = Entry(form, textvariable=password, width=120, font="Arial 13")
        passwd_entry.place(x=40, y=200)


        #Button
        signupbtn = Button(form, text="S I G N   U P", font="Arial 16 bold", bg=background, fg=frame, width=83, command=signup).place(x=40, y=250)



        self.root.mainloop()
MyGUI()